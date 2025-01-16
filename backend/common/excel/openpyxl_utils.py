# -*- coding: utf-8 -*-
"""
@Author  : yangkai
@Email   : 807440781@qq.com
@Project : Krun
@Module  : openpyxl_utils.py
@DateTime: 2025/1/16 12:38
"""
import os
import re
from typing import Optional, List, Dict, Union, Any, Generator, Literal, Tuple, Pattern

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class OpenpyxlUtils(object):
    """
    利用 openpyxl 库封装的Excel文件操作工具类
    :func create_excel:             如果目标文件路径上的excel文件不存在，则在该路径上创建excel实体文件
    :func create_sheet:             如果指定的sheet页名称不存在，则在该excel文件操作对象中新建sheet页
    :func acquire_excel_object:     如果目标文件路径上的excel文件存在，则获取该文件的操作对象
    :func acquire_sheet_object:     获取excel文件中指定的sheet页操作对象
    :func acquire_sheet_names:      获取excel文件中所有的sheet页名称
    :func acquire_sheet_cols:       获取sheet页中第一行所有的数据列内容（表头）和列数
    :func acquire_sheet_rows:       获取sheet页中所有的数据行、单元格坐标和行数
    :func acquire_sheet_data:       获取sheet页中所有的数据行和数据列中的内容，并根据表头内容压缩成列表嵌套字典对象，每个子元素代表一行数据
    :func acquire_cell_value:       获取sheet页中指定单元格的值
    :func acquire_sheet_series:     获取sheet页中指定行或列的所有单元格内容
    :func acquire_header_letter:    获取sheet页中指定的表头行对应的字母索引
    :func append_data_to_sheet:     在sheet页的尾部追加数据行或数据列，又或是，在指定行号或列号前插入数据行或数据列
    :func modify_cell_value:        修改sheet页中指定单元格的内容
    :func modify_series:            修改sheet页中指定数据行或数据列的所有单元格的内容
    :func delete_series:            删除sheet页中指定数据行或数据列的所有单元格的内容
    :func set_rows_height:          设置sheet页中所有数据行的高度，或者，设置指定行的高度
    :func set_cols_width:           设置sheet页中所有数据列的宽度（自适应单元格文本长度），或者，设置指定列的宽度
    :func set_cell_color:           设置指定单元格的背景颜色
    :func set_rows_color:           设置隔行变色（默认）将sheet页中偶数行设置背景颜色，或者，为指定行设置背景色
    :func set_cols_color:           设置隔列变色（默认）将sheet页中偶数列设置背景颜色，或者，为指定列设置背景色
    :func set_rows_style:           为指定的行，设置单元格样式
    :func set_cols_style:           为指定的列，设置单元格样式
    :func set_cell_hyperlink:       为指定的单元格，设置超链接
    :func set_cell_image:           为指定的单元格，插入图片
    :func set_range_cell_merged:    为指定的单元格区域设置单元格合并
    :func by_cell_value_set_color:  根据单元格的内容设置单元格背景色
    :func by_series_value_set_color:根据数据行或数据列的内容进行背景色填充
    :func by_header_acquire_letter: 根据表头名称获取字母索引、数字索引
    :func check_cell_is_merged:     检查指定坐标的单元格是否为合并单元格
    :func check_sheet_exists:       检查指定的sheet页名称是否存在
    :func save_excel:               保存文件
    """

    def __init__(self, file_path, read_only: bool = False):
        self.file_path = file_path
        self.read_only = read_only
        self.cell_style = {
            "实心填充": "solid",
            "居中对齐": "center", "左对齐": "left", "右对齐": "right",
            "红色": "DE5D83", "灰色": "E8E8E8",
            "黄色": "FFFF00", "白色": "FFFFFF",
            "蓝色": "3399FF", "粉色": "FFC0CB",
            "绿色": "90EE90", "紫色": "CC99FF",
            "细线": "thin", "双线": "double",
            "粗线": "thick", "无线": "none",

        }

    def create_excel(self, sheet_name: str = "Sheet1", sheet_data: Optional[List[list]] = None) -> bool:
        """
        如果目标文件路径上的excel文件不存在，则在该路径上创建excel实体文件
        :param sheet_name:  非必填项；字符类型；默认值：Sheet1；  创建excel文件时必须创建一个sheet页
        :param sheet_data:  非必填项；列表类型；默认值：None；    需要写入sheet页的数据，每个子列表代表一行数据
        :return: 新建成功返回True，新建失败返回False
        """
        # 目标文件路径已存在，跳过新建，返回False
        if os.path.exists(path=self.file_path):
            return False

        # 目标文件路径不存在，执行新建并修改sheet页名称
        excel_object: Workbook = Workbook()
        sheet_object: Worksheet = excel_object.active
        sheet_object.title = sheet_name

        # 检查是否需要写入数据，处理完成后保存文件，并返回True
        if sheet_data is not None:
            for row in sheet_data:
                sheet_object.append(row)
        excel_object.save(filename=self.file_path)
        return True

    def create_sheet(self, excel_object: Workbook, sheet_name: str, sheet_data: Optional[List[list]] = None,
                     header_name: Optional[list] = None) -> None:
        """
        如果指定的sheet页名称不存在，则在该excel文件操作对象中新建sheet页
        :param excel_object:    必填参数；对象类型；excel文件操作对象
        :param sheet_name:      必填参数；字符类型；目标新建sheet页名称
        :param sheet_data:      非必填项；列表类型；目标写入sheet页的数据内容
        :param header_name:     非必填项；列表类型；目标写入sheet页的表头内容
        :return:
        """
        if self.check_sheet_exists(excel_object=excel_object, sheet_name=sheet_name):
            raise ValueError(f"参数异常：无法新建同名sheet页")

        # 创建新的sheet页并为其命名
        excel_object.create_sheet(title=sheet_name)
        sheet_object = excel_object[sheet_name]
        # 检查是否需要写入数据
        if sheet_data is not None:
            for row in sheet_data:
                sheet_object.append(row)
        # 检查是否需要写入表头
        elif header_name is not None:
            for row in [header_name]:
                sheet_object.append(row)
        self.save_excel(excel_object=excel_object)

    def acquire_excel_object(self) -> Workbook:
        """
        如果目标文件路径上的Excel文件存在，则获取该文件的操作对象
        :return: Workbook
        """
        # 如果文件不存在，则抛出异常
        if not os.path.exists(path=self.file_path):
            raise FileNotFoundError(f"参数异常：文件路径不存在")

        excel_object: Workbook = openpyxl.load_workbook(
            filename=self.file_path,
            data_only=True,
            read_only=self.read_only
        )
        return excel_object

    def acquire_sheet_object(self, excel_object: Workbook, sheet_name: str) -> Worksheet:
        """
        获取excel文件中指定的sheet页操作对象
        :param excel_object:    必填参数；对象类型；excel文件操作对象
        :param sheet_name:      必填参数；字符类型；目标新建sheet页名称
        :return:
        """
        if sheet_name not in self.acquire_sheet_names(excel_object):
            raise KeyError(f"参数异常：上送sheet页名称不存在")

        sheet_object: Worksheet = excel_object[sheet_name]
        return sheet_object

    @staticmethod
    def acquire_sheet_names(excel_object: Workbook) -> List[str]:
        """
        获取excel文件中所有的sheet页名称
        :return:
        """
        sheet_names: List[str] = excel_object.sheetnames

        return sheet_names

    @staticmethod
    def acquire_sheet_cols(sheet_object: Worksheet) -> Dict[str, Union[list, int]]:
        """
        获取sheet页中第一行所有的数据列内容（表头）和列数
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :return: 以字典形式返回；
            {
                'data': ['xxx', 'xxx', 'xxx', 'xxx'],
                'count': x
            }
        """
        global index
        # 记录标题字段
        name = []
        # 使用enumerate有序遍历第一行的每一个字段
        for index, column in enumerate(list(sheet_object.rows)[0]):
            # 将当前字段添加到列表
            if column.value:
                name.append(column.value)
            else:
                break
        return {'data': name, 'count': index + 1}

    @staticmethod
    def acquire_sheet_rows(sheet_object: Worksheet, header_index: int = 0) -> Dict[str, Union[list, int]]:
        """
        获取sheet页中所有的数据行、单元格坐标和行数
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param header_index: 非必填项；数值类型；表头所在行
        :return: 以字典形式返回；
            {
                'data': [['xxx', 'xxx', 'xxx', 'xxx']],
                'local': [(<Cell 'xx'.xx>, <Cell 'xx'.xx>, <Cell 'xx'.xx>, <Cell 'xx'.xx>)],
                'count': x
            }
        """
        global index  # 记录除表头外总行数
        datagram = []  # 记录每行数据
        located = []  # 记录数据坐标
        # 使用enumerate有序遍历所有行
        for index, row in enumerate(sheet_object.rows):
            # 第一行数据为表头，跳过
            if index == header_index:
                continue
            # 当前行的第一个字段有值，则将记录当前行所有列数据和坐标（当某行第一个单元格内容为False时则会终止记录）
            # if row[0].value:
            #     datagram.append([cell.value for cell in row])
            #     located.append(row)
            # else:
            #     break

            datagram.append([cell.value for cell in row])
            located.append(row)
        return {'data': datagram, 'local': located, 'count': index}

    def acquire_sheet_data(self, excel_object: Workbook, sheet_name: str) -> List[Dict[str, Any]]:
        """
        获取sheet页中所有的数据行和数据列中的内容，并根据表头内容压缩成列表嵌套字典对象，每个子元素代表一行数据
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_name:   必填参数；字符类型；sheet页名称
        :return: 以列表套字典形式返回，每一个字典都是以表头和行数据组合成的Key-Value
            [{'XX': 'xx', 'XX': 'xx', 'XX': 'xx', 'XX': 'xx'}]
        """
        # 获取sheet对象
        sheet_object: Worksheet = self.acquire_sheet_object(
            excel_object=excel_object,
            sheet_name=sheet_name
        )
        # 获取sheet页中所有列的表头和列数
        cols = self.acquire_sheet_cols(sheet_object)
        cols_name, cols_num = cols.get("data"), cols.get("count")
        if not cols_name or not cols_num:
            raise KeyError(f"数据异常：获取表头或列数异常或为空")

        # 获取sheet页中所有行的坐标和总行数
        rows = self.acquire_sheet_rows(sheet_object)
        rows_local, rows_num = rows.get("local"), rows.get("count")
        if not rows_local or not rows_num:
            raise KeyError(f"数据异常：获取数据或行数异常或为空")

        # 压缩并转为字典：使用二维循环推导式（先看最右侧循环遍历每一行所使用的单元格，其次再在循环列的数量取出每一个单元格的值）
        return [dict(zip(cols_name, [row[i].value for i in range(cols_num)])) for row in rows_local]

    @staticmethod
    def acquire_cell_value(sheet_object: Worksheet, row_index: int, col_index: int) -> str:
        """
        获取sheet页中指定单元格的值
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param row_index:    必填参数；数值类型；行号
        :param col_index:    必填参数；数值类型；列号
        :return:
        """
        cell_value = sheet_object.cell(row=row_index, column=col_index)
        return cell_value

    def acquire_sheet_series(self, excel_object: Workbook, sheet_name: str, row_index: int = None,
                             col_index: int = None) -> list:
        """
        获取sheet页中指定行或列的所有单元格内容
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_name:   必填参数；字符类型；sheet页名称
        :param row_index:    非必填项；数值类型；行号
        :param col_index:    非必填项；数值类型；列号
        :return:
        """
        # 获取sheet页操作对象
        sheet_object: Worksheet = self.acquire_sheet_object(
            excel_object=excel_object,
            sheet_name=sheet_name
        )
        # 如果同时上送了行和列，获取行内容优先
        if row_index:
            row_data: Generator = sheet_object.iter_rows(
                min_row=row_index,
                max_row=row_index,
                values_only=True
            )
            row_data: tuple = row_data.__next__()
            return list(row_data)
        if col_index:
            col_data: Generator = sheet_object.iter_cols(
                min_col=col_index,
                max_col=col_index,
                values_only=True
            )
            col_data: tuple = col_data.__next__()
            return list(col_data)

    @staticmethod
    def acquire_header_letter(sheet_object: Worksheet, header_index: int = 1) -> Tuple[list, list]:
        """
        获取sheet页中指定的表头行对应的字母索引
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param header_index: 非必填项；数值类型；表头所在行
        :return:
        """
        header_names: list = []
        header_letters: list = []
        # 获取sheet页的表头对象
        sheet_header_name: list = sheet_object[header_index]
        for name in sheet_header_name:
            header_names.append(name.value)
            header_letters.append(name.column_letter)

        return header_names, header_letters

    def append_data_to_sheet(self, excel_object: Workbook, sheet_object: Worksheet,
                             append_data: List[list], append_name: Optional[List] = None,
                             append_index: Optional[int] = None, axis: Literal[0, 1] = 0) -> None:
        """
        在sheet页的尾部追加数据行或数据列，又或是，在指定行号或列号前插入数据行或数据列
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param append_data:  必填参数；列表类型；需要写入sheet页的数据，每一个子列表代表一行或一列
        :param append_name:  非必填项；列表类型；作用于列时使用，代表新建列的列名
        :param append_index: 非必填项；列表类型；插入时使用，默认是追加，代表在某行或某列前进行插入
        :param axis:         必填参数；数值类型；0表示垂直（添加行，默认），1表示水平（添加列）
        :return:
        """
        # 追加数据行（垂直）
        if axis == 0 and append_index is None:
            for row_data in append_data:
                sheet_object.append(row_data)
        # 插入数据行（垂直）
        elif axis == 0 and append_index is not None and isinstance(append_index, int):
            sheet_object.insert_rows(idx=append_index, amount=len(append_data))
            for row_index, row_data in enumerate(append_data, start=append_index):
                for col_index, col_data in enumerate(row_data, start=1):
                    sheet_object.cell(row=row_index, column=col_index, value=col_data)
        # 作用于数据列
        elif axis == 1:
            if not append_name or not append_data:
                raise ValueError("参数异常：axis为1时表示作用于数据列，append_name和append_data不可为None值")
            append_name_len = len(append_name)
            append_data_len = len(append_data)
            if append_name_len != append_data_len:
                raise ValueError("参数异常：axis为1时表示作用于数据列，append_name和append_data长度需要一致")

            # 插入数据列（水平）
            if append_index is not None and isinstance(append_index, int):
                append_start_column = append_index  # 从指定列数开始追加新的列
                append_ended_column = append_start_column + append_name_len
                sheet_object.insert_cols(idx=append_index, amount=append_name_len)
            # 追加数据行（水平）
            else:
                # 根据最大列数确定追加的位置
                append_start_column = sheet_object.max_column + 1  # 从最大列数之后开始追加新的列
                append_ended_column = append_start_column + append_data_len

            # 追加新的列名到工作表的标题行
            for col_num in range(append_start_column, append_ended_column):
                sheet_object.cell(row=1, column=col_num, value=append_name[col_num - append_start_column])

            # 追加新的数据到工作表的数据区域
            for col_num in range(append_start_column, append_ended_column):
                for row_num, row_data in enumerate(append_data[col_num - append_start_column]):
                    sheet_object.cell(
                        row=row_num + 2,
                        column=col_num,
                        value=row_data
                    )
        else:
            axis_msg: str = "axis为0时表示作用于数据行，axis为1时表示作用于数据列；为1时append_name不可为None值且需与append_data长度一致；"
            append_index_msg: str = "append_index不为None时表示在指定行索引或列索引处插入；"
            raise ValueError(f"参数异常：{axis_msg}{append_index_msg}")

        self.save_excel(excel_object=excel_object)

    def modify_cell_value(self, excel_object: Workbook, sheet_object: Worksheet, row_index: int, col_index: int,
                          value: str) -> None:
        """
        修改sheet页中指定单元格的内容
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param row_index:    必填参数；数值类型；行号
        :param col_index:    必填参数；数值类型；列号
        :param value:        必填参数；字符类型；写入单元格的内容
        :return:
        """
        sheet_object.cell(row=row_index, column=col_index, value=value)
        self.save_excel(excel_object=excel_object)

    def modify_series(self, excel_object: Workbook, sheet_object: Worksheet, index: int, new_values: list,
                      axis: Literal[0, 1]) -> None:
        """
        修改sheet页中指定数据行或数据列的所有单元格的内容
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param index:        必填参数；数值类型；行号或列号
        :param new_values:   必填参数；列表类型；写入某数据行或数据列的数据内容
        :param axis:         必填参数；数值类型；0表示垂直（行），1表示水平（列）
        :return:
        """
        new_value_len = len(new_values)
        if axis == 0:
            for col in range(1, new_value_len + 1):
                sheet_object.cell(row=index, column=col, value=new_values[col - 1])
        elif axis == 1:
            for row in range(1, new_value_len + 1):
                sheet_object.cell(row=row + 1, column=index, value=new_values[row - 1])
        self.save_excel(excel_object=excel_object)

    def delete_series(self, excel_object: Workbook, sheet_object: Worksheet, axis: Literal[0, 1],
                      index: int, count: int = 1) -> None:
        """
        删除sheet页中指定数据行或数据列的所有单元格的内容
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param axis:         必填参数；数值类型；0表示垂直（行），1表示水平（列）
        :param index:        必填参数；数值类型；行号或列号
        :param count:        非必填项，数值类型；删除数量，默认删除1行或1列
        :return:
        """
        if axis == 0:
            sheet_object.delete_rows(idx=index, amount=count)
        elif axis == 1:
            sheet_object.delete_cols(idx=index, amount=count)
        else:
            raise ValueError("参数异常：axis参数只接受0或1")

        self.save_excel(excel_object=excel_object)

    def set_rows_height(self, excel_object: Workbook, sheet_object: Worksheet,
                        rows: Optional[list] = None, row_height: int = 30):
        """
        设置sheet页中所有数据行的高度，或者，设置指定行的高度
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param rows:         非必填项；列表类型；需要指定设置行高的行号
        :param row_height:   非必填项；数值类型；设置数据行的高度（默认30）
        :return:
        """
        if rows is None:
            # 循环设置每一行的高度
            max_row = sheet_object.max_row
            for row in range(1, max_row + 1):
                sheet_object.row_dimensions[row].height = row_height
        else:
            for row in rows:
                sheet_object.row_dimensions[row].height = row_height
        self.save_excel(excel_object=excel_object)

    def set_cols_width(self, excel_object: Workbook, sheet_object: Worksheet,
                       auto_width: bool = True, cols: Optional[list] = None,
                       col_width: int = 30):
        """
        设置sheet页中所有数据列的宽度（自适应单元格文本长度），或者，设置指定列的宽度
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param auto_width:   非必填项；布尔类型；默认为True；为所有列设置自适应宽度
        :param cols:         非必填项；列表类型；指定需要设置列宽的列号
        :param col_width:    非必填项；数值类型；设置指定列号宽度（默认30）
        :return:
        """

        def count_chinese_char(string: str) -> int:
            # 预编译一个正则表达式，匹配所有中文
            _pattern: str = r"[\u4e00-\u9fa5]"
            _compile: Pattern = re.compile(_pattern)
            # 使用findall找出所有匹配正则表达式的结果
            _matches: List[str] = _compile.findall(string)
            return len(_matches)

        if auto_width:
            # 获取当前sheet页的数据shape
            max_row, max_column = sheet_object.max_row, sheet_object.max_column
            for col in range(max_column):
                # 根据列的数字返回对应的字母，暂定每列宽度为1
                max_widths, col_name = 1, get_column_letter(col + 1)
                # 获取当前列下的所有数据，取单元格值内数据最大长度
                for row in range(max_row):
                    cell_value = str(sheet_object[f'{col_name}{row + 1}'].value)
                    chinese_count = count_chinese_char(cell_value)
                    content_length = len(str(cell_value or '' or None)) + chinese_count + 2
                    # 如果单元格内最大长度比假设大则修改
                    max_widths = content_length if content_length > max_widths else max_widths

                # 把当前列所有单元格的长度设为最大长度
                max_widths = 100 if max_widths > 100 else max_widths
                sheet_object.column_dimensions[col_name].width = max_widths
        else:
            for col in cols:
                column_letter = get_column_letter(col)
                if column_letter not in sheet_object.column_dimensions:
                    sheet_object.column_dimensions[column_letter] = sheet_object.column_dimensions[column_letter]
                else:
                    del sheet_object.column_dimensions[column_letter]
                    sheet_object.column_dimensions[column_letter] = sheet_object.column_dimensions[column_letter]
                sheet_object.column_dimensions[column_letter].width = col_width

        self.save_excel(excel_object=excel_object)

    def set_cell_color(self, excel_object: Workbook, sheet_object: Worksheet, row: int, col: int,
                       color_name: Literal["红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"]) -> None:
        """
        设置指定单元格的背景颜色
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param row:          必填参数；数值类型；行号
        :param col:          必填参数；数值类型；列号
        :param color_name:   必填参数；字符类型；预设的颜色名称
        :return:
        """
        if not row or not col:
            raise ValueError("参数异常：row和col必须大于等于1")

        color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(color_name)
        )
        sheet_object.cell(row=row, column=col).fill = color

        self.save_excel(excel_object=excel_object)

    def set_rows_color(self, excel_object: Workbook, sheet_object: Worksheet, rows: Optional[list] = None,
                       color_name: Literal["红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"] = "灰色") -> None:
        """
        设置隔行变色（默认）将sheet页中偶数行设置背景颜色，或者，为指定行设置背景色
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param rows:         非必填项；列表类型；需要指定设置背景色的行号
        :param color_name:   非必填项；字符类型；预设的颜色名称
        :return:
        """
        max_row, max_col = sheet_object.max_row, sheet_object.max_column
        color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(color_name)
        )
        if rows is None:
            for row in range(1, max_row + 1):
                # 如果i能被2整除就可以得到偶数行
                if row % 2 == 0:
                    # 循环设置每一行的背景颜色
                    for col in range(1, max_col + 1):
                        sheet_object.cell(row=row, column=col).fill = color
        else:
            for row in rows:
                for col in range(1, max_col + 1):
                    sheet_object.cell(row=row, column=col).fill = color

        self.save_excel(excel_object=excel_object)

    def set_cols_color(self, excel_object: Workbook, sheet_object: Worksheet, cols: Optional[list] = None,
                       color_name: Literal["红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"] = "灰色",
                       start_index: int = 2) -> None:
        """
        设置隔列变色（默认）将sheet页中偶数列设置背景颜色，或者，为指定列设置背景色
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param cols:         非必填项；列表类型；需要指定设置背景色的行号
        :param color_name:   非必填项；字符类型；预设的颜色名称
        :param start_index:  非必填项；数值类型；表头所在行（是否作用到表头）
        :return:
        """
        max_row, max_col = sheet_object.max_row, sheet_object.max_column
        color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(color_name)
        )
        if cols is None:
            for col in range(1, max_col + 1):
                # 如果i能被2整除就可以得到偶数列
                if col % 2 == 0:
                    # 循环设置偶数列当前单元格的背景颜色
                    for row in range(start_index, max_row + 1):
                        sheet_object.cell(row=row, column=col).fill = color
        else:
            for col in cols:
                # 循环设置偶数列当前单元格的背景颜色
                for row in range(start_index, max_row + 1):
                    sheet_object.cell(row=row, column=col).fill = color

        self.save_excel(excel_object=excel_object)

    def set_rows_style(self, excel_object: Workbook, sheet_object: Worksheet,
                       start_index: int = 2, rows: Optional[list] = None,
                       is_bold: bool = True, is_border: bool = True,
                       alignment_type: Literal["居中对齐", "左对齐", "右对齐"] = "居中对齐") -> None:
        """
        为指定的行，设置单元格样式
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param start_index:  非必填项；数值类型；表头所在行（是否作用到表头）
        :param rows:         非必填项；列表类型；需要指定设置样式的行号
        :param is_bold:      必填参数；布尔类型；是否加粗
        :param is_border:    必填参数；布尔类型；是否加边框
        :param alignment_type:必填参数；字符类型；预设的对其方式名称
        :return:
       """
        side = Side(
            style=self.cell_style.get("细线"),
            color=self.cell_style.get("黑色")
        )
        max_row, max_col = sheet_object.max_row, sheet_object.max_column

        condition: dict = {
            "font": Font(bold=True),
            "border": Border(left=side, right=side, top=side, bottom=side),
            "alignment": Alignment(
                horizontal=self.cell_style.get(alignment_type),
                vertical=self.cell_style.get(alignment_type)
            )
        }
        if rows is None:
            rows = range(start_index, max_row + 1)
        else:
            if not is_bold: condition.pop("font")
            if not is_border: condition.pop("border")

        for row in rows:
            for col in range(1, max_col + 1):
                sheet_object.cell(row=row, column=col).font = condition.get("font")
                sheet_object.cell(row=row, column=col).border = condition.get("border")
                sheet_object.cell(row=row, column=col).alignment = condition.get("alignment")
        self.save_excel(excel_object=excel_object)

    def set_cols_style(self, excel_object: Workbook, sheet_object: Worksheet,
                       start_index: int = 2,
                       cols: Optional[list] = None,
                       is_bold: bool = True,
                       is_border: bool = True,
                       alignment_type: Literal["居中对齐", "左对齐", "右对齐"] = "居中对齐") -> None:
        """
        为指定的列，设置单元格样式
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param start_index:  非必填项；数值类型；表头所在行（是否作用到表头）
        :param cols:         非必填项；列表类型；需要指定设置样式的列号
        :param is_bold:      必填参数；布尔类型；是否加粗
        :param is_border:    必填参数；布尔类型；是否加边框
        :param alignment_type:必填参数；字符类型；预设的对其方式名称
        :return:
       """
        side = Side(
            style=self.cell_style.get("细线"),
            color=self.cell_style.get("黑色")
        )
        max_row, max_col = sheet_object.max_row, sheet_object.max_column

        condition: dict = {
            "font": Font(bold=True),
            "border": Border(left=side, right=side, top=side, bottom=side),
            "alignment": Alignment(
                horizontal=self.cell_style.get(alignment_type),
                vertical=self.cell_style.get(alignment_type)
            )
        }
        if cols is None:
            cols = range(start_index, max_row + 1)
        else:
            if not is_bold: condition.pop("font")
            if not is_border: condition.pop("border")

        for col in cols:
            for row in range(start_index, max_row + 1):
                sheet_object.cell(row=row, column=col).font = condition.get("font")
                sheet_object.cell(row=row, column=col).border = condition.get("border")
                sheet_object.cell(row=row, column=col).alignment = condition.get("alignment")
        self.save_excel(excel_object=excel_object)

    def set_cell_hyperlink(self, excel_object: Workbook, sheet_object: Worksheet,
                           rows: List[int], cols: List[int], link_paths: List[str],
                           link_names: Optional[List[str]] = None,
                           color_name: Literal[
                               "红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"] = "蓝色") -> None:
        """
        为指定的单元格，设置超链接
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param rows:         必填参数；列表类型；需要指定设置超链接的行号
        :param cols:         必填参数；列表类型；需要指定设置超链接的列号
        :param link_paths:   必填参数；列表类型；需要指定设置超链接的路径
        :param link_names:   非必填项；列表类型；需要指定设置超链接的显示文本
        :param color_name:   非必填项；字符类型；预设的颜色名称
        :return:
       """
        rows_len = len(rows)
        cols_len = len(cols)
        link_paths_len = len(link_paths)
        if not (rows_len == cols_len == link_paths_len):
            raise ValueError("参数异常：rows、cols、link_paths的长度必须一致")
        if link_names is not None and len(link_names) != link_paths_len:
            raise ValueError("参数异常：当link_names不为None时，必须和link_paths参数长度一致")
        if link_names is None:
            link_names = ["超链接" + str(num) for num in range(1, link_paths_len + 1)]

        color = Font(bold=True, color=self.cell_style.get(color_name))
        for row, col, path, name in zip(rows, cols, link_paths, link_names):
            sheet_object.cell(row=row, column=col).value = f'=HYPERLINK("{path}","{name}")'
            sheet_object.cell(row=row, column=col).font = color

        self.save_excel(excel_object=excel_object)

    def set_cell_image(self, excel_object: Workbook, sheet_object: Worksheet,
                       rows: List[int], cols: List[int], image_paths: List[str],
                       row_height: int = 50, col_width: int = 35) -> None:
        """
        为指定的单元格，插入图片
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param rows:         必填参数；列表类型；需要插入图片的行号
        :param cols:         必填参数；列表类型；需要插入图片的列号
        :param image_paths:  必填参数；列表类型；需要插入图片的路径
        :param row_height:   非必填项；数值类型；设置单元格的高度（作用到图片）
        :param col_width:    非必填项；数值类型；设置单元格的宽度（作用到图片）
        :return:
       """
        rows_len = len(rows)
        cols_len = len(cols)
        image_paths_len = len(image_paths)
        if not (rows_len == cols_len == image_paths_len):
            raise ValueError("参数异常：rows、cols、image_paths的长度必须一致")

        for row, col, path in zip(rows, cols, image_paths):
            image = Image(path)
            column_letter = get_column_letter(col)

            image.width, image.height = col_width * 8, row_height * 1.36
            sheet_object.add_image(image, f"{column_letter}{row}")
            sheet_object.row_dimensions[row].height = row_height
            if column_letter not in sheet_object.column_dimensions:
                sheet_object.column_dimensions[column_letter] = sheet_object.column_dimensions[column_letter]
            else:
                del sheet_object.column_dimensions[column_letter]
                sheet_object.column_dimensions[column_letter] = sheet_object.column_dimensions[column_letter]

            sheet_object.column_dimensions[column_letter].width = col_width
        self.save_excel(excel_object=excel_object)

    def set_range_cell_merged(self, excel_object: Workbook, sheet_object: Worksheet,
                              shape: List[List[int]]) -> None:
        """
        为指定的单元格区域设置单元格合并
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param shape: 二维列表，[开始行、开始列、结束行、结束列]
        :return:
        """
        for _range in shape:
            left_up_anchor = _range[0]
            left_down_anchor = _range[1]
            right_up_anchor = _range[2]
            right_down_anchor = _range[3]
            sheet_object.merge_cells(
                start_row=left_up_anchor,
                start_column=left_down_anchor,
                end_row=right_up_anchor,
                end_column=right_down_anchor
            )
        self.save_excel(excel_object=excel_object)

    def by_cell_value_set_color(self, excel_object: Workbook, sheet_object: Worksheet, row: int, col: int,
                                color_name: Literal["红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"],
                                expected_value: Union[int, str, None] = None) -> None:
        """
        根据单元格的内容设置单元格背景色
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param row:          必填参数；数值类型；行号
        :param col:          必填参数；数值类型；列号
        :param color_name:   必填参数；字符类型；预设的颜色名称
        :param expected_value:非必填项；字符类型；预期值，如果上送该参数，则会判断是否与单元格上的内容一致
        :return:
        """
        if not row or not col:
            raise ValueError("参数异常：row和col必须大于等于1")

        color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(color_name)
        )
        if expected_value is None or sheet_object.cell(row=row, column=col).value == expected_value:
            sheet_object.cell(row=row, column=col).fill = color

        self.save_excel(excel_object=excel_object)

    def by_series_value_set_color(self, excel_object: Workbook, sheet_object: Worksheet, expected_values: list,
                                  axis: Literal[0, 1] = 0, rows: Optional[list] = None, cols: Optional[list] = None,
                                  start_index: int = 2,
                                  passed_color_name: Literal[
                                      "红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"] = "绿色",
                                  failed_color_name: Literal[
                                      "红色", "黄色", "蓝色", "绿色", "灰色", "白色", "粉色", "紫色"] = "红色"
                                  ):
        """
        根据数据行或数据列的内容进行背景色填充
        :param excel_object:    必填参数；对象类型；excel文件操作对象
        :param sheet_object:    必填参数；对象类型；excel文件中的sheet页操作对象
        :param expected_values: 必填参数；字符类型；预期值，如果上送该参数，则会判断是否与单元格上的内容一致
        :param axis:            非必填项；数值类型；0表示垂直（行）（默认），1表示水平（列）
        :param rows:            非必填项；列表类型；需要根据单元格的内容进行填充背景色的行号
        :param cols:            非必填项；列表类型；需要根据单元格的内容进行填充背景色的列号
        :param start_index:     非必填项；数值类型；表头所在行（是否作用到表头）
        :param passed_color_name:非必填项；字符类型；预设的颜色名称作用于匹配预期值
        :param failed_color_name:非必填项；字符类型；预设的颜色名称作用于不匹配预期值
        :return:
        """
        if axis == 0 and not rows:
            raise ValueError("参数异常：axis为0时表示作用于数据行，rows不可为None值")

        if axis == 1 and not cols:
            raise ValueError("参数异常：axis为1时表示作用于数据列，cols不可为None值")

        max_row, max_col = sheet_object.max_row, sheet_object.max_column
        passed_color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(passed_color_name)
        )
        failed_color = PatternFill(
            patternType=self.cell_style.get("实心填充"),
            fgColor=self.cell_style.get(failed_color_name)
        )
        if axis == 0:
            for row in rows:
                for col in range(1, max_col + 1):
                    if sheet_object.cell(row=row, column=col).value in expected_values:
                        sheet_object.cell(row=row, column=col).fill = passed_color
                    else:
                        sheet_object.cell(row=row, column=col).fill = failed_color
        else:
            for col in cols:
                for row in range(start_index, max_row + 1):
                    if sheet_object.cell(row=row, column=col).value in expected_values:
                        sheet_object.cell(row=row, column=col).fill = passed_color
                    else:
                        sheet_object.cell(row=row, column=col).fill = failed_color
        self.save_excel(excel_object=excel_object)

    @staticmethod
    def by_header_acquire_letter(sheet_object, header_names: list, header_index: int = 1) -> Tuple[list, list, list]:
        """
        根据表头名称获取字母索引、数字索引
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param header_names: 必填参数；列表类型；表头内容
        :param header_index: 必填参数；列表类型；表头内容所在行
        :return:
        """
        column_letters = []
        column_indexes = []
        sheet_header_cells: tuple = sheet_object[header_index]
        for header_cell in sheet_header_cells:
            for header_name in header_names:
                if header_cell.value == header_name:
                    column_letter = header_cell.column_letter
                    column_letters.append(column_letter)
                    column_indexes.append(column_index_from_string(column_letter))
        return header_names, column_letters, column_indexes

    @staticmethod
    def check_cell_is_merged(sheet_object: Worksheet,
                             row_index: int, col_index: int) -> Tuple[bool, any]:
        """
        检查指定坐标的单元格是否为合并单元格
        :param sheet_object: 必填参数；对象类型；excel文件中的sheet页操作对象
        :param row_index:    必填参数；数值类型；需要检查是否为合并单元格的行号
        :param col_index:    必填参数；数值类型；需要检查是否为合并单元格的列号
        :return:
        """
        all_merged_cells = sheet_object.merged_cell_ranges
        for merged_cell in all_merged_cells:
            min_row = merged_cell.min_row  # 合并单元格最小行
            max_row = merged_cell.max_row  # 合并单元格最大行
            min_col = merged_cell.min_col  # 合并单元格最小列
            max_col = merged_cell.max_col  # 合并单元格最大列
            if (min_row <= row_index <= max_row) and (min_col <= col_index <= max_col):
                merged_cell_value = sheet_object.cell(min_row, min_col).value
                return True, merged_cell_value
        return False, None

    def check_sheet_exists(self, excel_object: Workbook, sheet_name: str) -> bool:
        """
        检查指定的sheet页名称是否存在
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :param sheet_name:   必填参数；字符类型；sheet页名称
        :return:
        """
        return sheet_name in self.acquire_sheet_names(excel_object=excel_object)

    def save_excel(self, excel_object: Workbook):
        """
        检查指定的sheet页名称是否存在
        :param excel_object: 必填参数；对象类型；excel文件操作对象
        :return:
        """
        excel_object.save(self.file_path)
