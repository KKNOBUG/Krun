# -*- coding: utf-8 -*-
"""
@Author  : yangkai
@Email   : 807440781@qq.com
@Project : Krun
@Module  : pandas_utils.py
@DateTime: 2025/1/16 12:40
"""

import os

import pandas as pd
from pandas import DataFrame
from typing import Literal, List, Optional


class PandasUtils(object):
    """
    利用 pandas 库封装的Excel文件操作工具类
    :func create_excel:  如果目标文件路径上的excel文件不存在，则在该路径上创建excel实体文件
    :func create_sheet:  在文件中尝试创建一个新的sheet页数据
    :func read_sheet_data:  获取指定sheet页的数据视图
    :func acquire_sheet_data:  获取指定sheet页的数据视图，并按照指定的表头，压缩成列表嵌套字典，列表中每一个子元素代表一行数据
    :func acquire_cell_value:  获取指定sheet页的数据视图中的指定行列坐标的数据
    :func acquire_sheet_series:  获取指定sheet页数据视图中的一列或一行数据
    :func acquire_sheet_names:  获取excel文件中所有的sheet页名称
    :func check_sheet_exists:  检查指定sheet页名称是否存在
    :func append_data_to_sheet:  读取原文件中的指定sheet页，转换为dataframe数据视图，在其后追加新的数据行（不限制长度）或者新的数据列（长度不可超过原始列）
    :func modify_cell_value:  读取原文件中的指定sheet页，转换为dataframe数据视图，指定坐标（行号+列号）的单元格内容
    :func modify_series:  读取原文件中的指定sheet页，转换为dataframe数据视图，修改其中的行（axis=1）或者列（axis=0）
    :func filter_sheet_data:  过滤、匹配指定sheet页中的数据（in、not in、==、!=、>、>=、<、<=）
    """

    def __init__(self, file_path) -> None:
        """
        PandasExcelTool工具类构造方法
        :param file_path: 必填参数；字符串类型；目标文件路径
        """
        self.file_path = file_path

    def create_excel(self, sheet_name: str = "Sheet1", sheet_data: Optional[List[list]] = None) -> bool:
        """
        如果目标文件路径上的excel文件不存在，则在该路径上创建excel实体文件
        :param sheet_name:  非必填项；字符类型；默认值：Sheet1；  创建excel文件时必须创建一个sheet页
        :param sheet_data:  非必填项；列表类型；默认值：None；    需要写入sheet页的数据，每个子列表代表一行数据
        :return: 新建成功返回True，新建失败返回False
        """
        if os.path.exists(path=self.file_path):
            return False

        # 创建ExcelWriter对象
        with pd.ExcelWriter(path=self.file_path, engine='xlsxwriter') as writer:
            if sheet_data is not None:
                # 如果参数sheet_data不为None则说明有数据写入sheet页，将之转换成数据视图
                data_frame = pd.DataFrame(data=sheet_data)
            else:
                # 创建一个空的DataFrame作为工作表
                data_frame = pd.DataFrame()
            # 将DataFrame写入Excel文件的工作表中
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            return True

    def create_sheet(self, sheet_name: str, sheet_data: Optional[List[list]] = None,
                     header_columns: Optional[list] = None) -> None:
        """
        在文件中尝试创建一个新的sheet页数据
        如果sheet已经存在则抛出异常
        如果sheet_data不为None，则新建一个sheet页并写入sheet_data数据
        如果header_columns不为None，则新建一个sheet页并写入表头内容
        :param sheet_name: excel文件中的sheet名称
        :param sheet_data: 需要写入sheet页的数据
        :param header_columns: 表头内容
        :return:
        """
        if sheet_name in self.acquire_sheet_names():
            raise ValueError(f"sheet {sheet_name} 已经存在，无法新建同名sheet页")

        # 打开Excel文件
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a') as writer:
            # 在Excel文件中创建一个新的sheet页
            if sheet_data:
                new_sheet = pd.DataFrame(sheet_data)
                new_sheet.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            else:
                new_sheet = pd.DataFrame(columns=header_columns) if header_columns else pd.DataFrame()
                new_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    def read_sheet_data(self, sheet_name, index_col=None, header=None) -> DataFrame:
        """
        获取指定sheet页的数据视图
        :param sheet_name: excel文件中的sheet名称
        :param index_col:  从指定sheet页中的哪一行开始读取
        :param header:     设置表头索引
        :return:
        """
        return pd.read_excel(io=self.file_path, sheet_name=sheet_name, index_col=index_col, header=header)

    def acquire_sheet_data(self, sheet_name: str, header_row_index: int = 0) -> List[dict]:
        """
        获取指定sheet页的数据视图，并按照指定的表头，压缩成列表嵌套字典，列表中每一个子元素代表一行数据
        :param sheet_name: excel文件中的sheet名称
        :param header_row_index: 表头行号
        :return:
        """
        data = pd.read_excel(self.file_path, sheet_name, header=header_row_index)
        return data.to_dict(orient='records')

    def acquire_cell_value(self, sheet_name: str, row_index: int = None, col_index: int = None,
                           index_col: int = None, header: int = None) -> any:
        """
        获取指定sheet页的数据视图中的指定行列坐标的数据
        :param sheet_name:
        :param row_index:
        :param col_index:
        :param index_col:
        :param header:
        :return:
        """
        df = self.read_sheet_data(sheet_name, index_col=index_col, header=header)
        cell_value: any = df.loc[row_index, col_index]
        return cell_value

    def acquire_sheet_series(self, sheet_name: str, row_index: int = None, col_index: int = None,
                             index_col: int = None, header: int = None) -> list:
        """
        获取指定sheet页数据视图中的一列或一行数据
        :param sheet_name:  excel文件中的sheet名称
        :param row_index:   行号
        :param col_index:   列号
        :param index_col:   从指定sheet页中的哪一行开始读取
        :param header:      设置表头索引
        :return:
        """
        df = self.read_sheet_data(sheet_name, index_col=index_col, header=header)
        if row_index is not None:
            row_data = df.iloc[row_index].tolist()
            return row_data
        if col_index is not None:
            col_data = df.loc[:, col_index].tolist()
            return col_data

    def acquire_sheet_names(self):
        """
        获取excel文件中所有的sheet页名称
        :return:
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"文件路径异常或不存在：「{self.file_path}」")
        excel_object = pd.ExcelFile(self.file_path, engine="openpyxl")
        sheet_names = excel_object.sheet_names
        return sheet_names

    def check_sheet_exists(self, sheet_name):
        """
        检查指定sheet页名称是否存在
        :param sheet_name:
        :return:
        """
        return sheet_name in self.acquire_sheet_names()

    def append_data_to_sheet(self, sheet_name: str, append_data: Optional[List[list]],
                             append_name: Optional[List] = None, axis: Literal[0, 1] = 0) -> None:
        """
        读取原文件中的指定sheet页，转换为dataframe数据视图，在其后追加新的数据行（不限制长度）或者新的数据列（长度不可超过原始列）
        :param sheet_name:  必填参数，字符类型，excel文件中的sheet名称
        :param append_data: 必填参数，列表类型，新列数据内容（二位数组）每一个子列表代表一行或一列
        :param append_name: 非必填项，列表类型，新列名称（如果是作用于列，需要传递该参数，并且需与append_data长度一致）
        :param axis:        axis=0（垂直拼接）作用于行；axis=1（水平拼接）作用于列；
        :return:
        """
        if not self.check_sheet_exists(sheet_name):
            raise KeyError(f"指定sheet名称「{sheet_name}」不存在")

        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='r+', if_sheet_exists="overlay") as writer:
            source_data_frame = pd.read_excel(self.file_path, sheet_name=sheet_name)
            if axis == 0:
                append_row_index = source_data_frame.shape[0] + 1
                append_data_frame = pd.DataFrame(append_data, index=None, columns=None)
                append_data_frame.to_excel(writer, sheet_name, startrow=append_row_index, index=False, header=False)
            else:
                append_name_len = len(append_name)
                append_data_len = len(append_data)
                if append_name_len != append_data_len:
                    raise ValueError(f"新增数据列名长度「{append_name_len}」与新增列数据长度「{append_data_len}」不一致")
                append_data_frame = self.__concat_column(source_data_frame, append_name, append_data)
                append_data_frame.to_excel(writer, sheet_name, index=False, header=True)

    def modify_cell_value(self, sheet_name: str, row_index: int, col_index: int, value: str):
        """
        读取原文件中的指定sheet页，转换为dataframe数据视图，指定坐标（行号+列号）的单元格内容
        :param sheet_name: excel文件中的sheet名称
        :param row_index:  行号
        :param col_index:  列号
        :param value:      单元格内容
        :return:
        """
        import openpyxl
        from openpyxl.workbook import Workbook
        excel_object: Workbook = openpyxl.load_workbook(self.file_path)
        sheet = excel_object[sheet_name]
        sheet.cell(row=row_index + 1, column=col_index + 1, value=value)
        excel_object.save(self.file_path)

    def modify_series(self, sheet_name, index: int, new_values: list, axis: Literal[0, 1]):
        """
        读取原文件中的指定sheet页，转换为dataframe数据视图，修改其中的行（axis=1）或者列（axis=0）
        注意：新数据的长度不可超过原数据视图的行或列长度
        :param sheet_name: excel文件中的sheet名称
        :param index:      行索引或者列索引
        :param new_values: 行数据或者列数据
        :param axis:       作用于行或者作用于列
        :return:
        """
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode="r+", if_sheet_exists="overlay") as writer:
            df = pd.read_excel(self.file_path, sheet_name, engine="openpyxl")
            row_count, col_count = df.shape
            if axis == 0:
                if len(new_values) != row_count:
                    raise ValueError(f"新数据长度「{len(new_values)}」，原视图列长度「{row_count}」")
                df.iloc[:, index] = new_values
            else:
                if len(new_values) != col_count:
                    raise ValueError(f"新数据长度「{len(new_values)}」，原视图行长度「{col_count}」")
                df.iloc[index] = new_values

            for column_name in df.columns:
                if str(column_name).startswith("Unnamed"):
                    df = df.rename(columns={column_name: None})
            df.to_excel(writer, sheet_name, index=False)

    def filter_sheet_data(self, sheet_name, conditions: List[tuple], index_col=None, header=None) -> pd.DataFrame:
        """
        过滤、匹配指定sheet页中的数据（in、not in、==、!=、>、>=、<、<=）
        conditions = [
            # ("薪资", "in", [3, 4], "&"),
            # ("薪资", "not in", [1, 2, 3], "&"),
            ("工作", "==", 2, "&"),
            # ("薪资", "!=", 1, "&"),
            ("薪资", ">", 1, "&"),
            # ("薪资", ">=", 1, "&"),
            # ("薪资", "<", 4, "&"),
            # ("薪资", "<=", 4, "&"),
        ]
        :param sheet_name:
        :param conditions:
        :param index_col:
        :param header:
        :return:
        """
        df = self.read_sheet_data(sheet_name, index_col=index_col, header=header)
        query_conditions = ""
        for col, op, val, sp in conditions:
            if op == "in":
                if not isinstance(val, list): raise TypeError("运算符 in 期待的是列表类型")
                query_conditions += f"{sp} (df['{col}'].isin({val})) "
            elif op == "not in":
                if not isinstance(val, list): raise TypeError("运算符 not in 期待的是列表类型")
                query_conditions += f"{sp} (~df['{col}'].isin({val})) "
            else:
                if isinstance(val, list): raise TypeError(f"运算符 {op} 期待的是非容器类型")
                query_conditions += f"{sp} (df['{col}'] {op} {val}) "

        query_conditions = "df[ " + query_conditions[2:] + "]"
        filtered_data = eval(query_conditions)

        return filtered_data

    @staticmethod
    def __concat_column(dataframe: DataFrame, column_name: list, column_data: list):
        for col_name, col_data in zip(column_name, column_data):
            # 插入loc指定位置，如指定位置存在数据则将之后移（此时设置为dataframe的最后一列），允许列名重复
            dataframe.insert(column=col_name, value=col_data, loc=dataframe.shape[1], allow_duplicates=True)
            # dataframe[col_name] = col_data  # 拼接在dataframe数据视图的最后，不允许列名重复
        return dataframe
